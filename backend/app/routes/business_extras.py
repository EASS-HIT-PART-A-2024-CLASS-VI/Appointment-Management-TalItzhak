from fastapi import APIRouter, HTTPException, Depends
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from app.schemas import AppointmentResponse
from app.models import Appointment, User, Service
from app.dependencies import get_db, check_user_role, get_current_user
from sqlalchemy import and_, text
from typing import List
import pandas as pd
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
import io
from datetime import datetime, timedelta

router = APIRouter()

# Business owners only access
business_owner_required = check_user_role("business_owner")

def check_business_ownership(business_id: int, current_user: dict, db: Session):
    """Helper function to check if the current user owns the business"""
    user = db.query(User).filter(User.username == current_user["sub"]).first()
    if not user or user.id != business_id:
        raise HTTPException(
            status_code=403,
            detail="You don't have permission to access this business's data"
        )
    return user


@router.get("/appointments/export")
async def export_appointments_to_excel(
    db: Session = Depends(get_db),
    current_user: dict = Depends(business_owner_required)
):
    try:
        print("Starting export process...")
        # Get the business owner
        user = db.query(User).filter(User.username == current_user["sub"]).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        # Get owner's services
        services = db.query(Service).filter(Service.owner_id == user.id).all()
        service_names = [service.name for service in services]

        # Get appointments for these services
        appointments = db.query(Appointment).filter(
            Appointment.type.in_(service_names)
        ).order_by(Appointment.date, Appointment.start_time).all()

        # Create DataFrame
        data = []
        for apt in appointments:
            data.append({
                'Date': apt.date.strftime('%Y-%m-%d') if apt.date else '',
                'Time': apt.start_time.strftime('%H:%M') if apt.start_time else '',
                'Service': apt.type or '',
                'Duration (minutes)': apt.duration or 0,
                'Customer Name': apt.customer_name or '',
                'Customer Phone': apt.customer_phone or '',
                'Cost': apt.cost or 0,
                'Notes': apt.notes or ''
            })

        # Create DataFrame with default columns if no data
        if not data:
            data = [{
                'Date': '',
                'Time': '',
                'Service': '',
                'Duration (minutes)': 0,
                'Customer Name': '',
                'Customer Phone': '',
                'Cost': 0,
                'Notes': ''
            }]

        df = pd.DataFrame(data)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Appointments')

        output.seek(0)

        # Use StreamingResponse instead of Response
        return StreamingResponse(
            BytesIO(output.getvalue()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=appointments_{datetime.now().strftime("%Y%m%d")}.xlsx'
            }
        )

    except Exception as e:
        print(f"Export error: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )

@router.get("/appointments/{appointment_id}")
def get_appointment_details(
    appointment_id: int,
    business_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(business_owner_required)
):
    # Check business ownership
    check_business_ownership(business_id, current_user, db)
    
    # Get appointment
    appointment = db.query(Appointment).filter(
        Appointment.id == appointment_id
    ).first()
    
    if not appointment:
        raise HTTPException(status_code=404, detail="Appointment not found")
        
    return appointment

@router.get("/appointments")
def list_appointments(
    business_id: int,
    title: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(business_owner_required)
):
    # Check business ownership
    check_business_ownership(business_id, current_user, db)
    
    # Query appointments
    query = db.query(Appointment)
    if title:
        query = query.filter(Appointment.title == title)
    
    return query.all()

@router.get("/appointments/search/{phone}")
def search_appointment(
   phone: str,
   db: Session = Depends(get_db),
   current_user: dict = Depends(business_owner_required)
):
   # Get the business owner
   user = db.query(User).filter(User.username == current_user["sub"]).first()
   if not user:
       raise HTTPException(status_code=404, detail="User not found")

   # Get the business owner's services
   services = db.query(Service).filter(Service.owner_id == user.id).all()
   service_names = [service.name for service in services]

   # Search appointments matching phone and business services
   appointments = db.query(Appointment).filter(
       Appointment.customer_phone.contains(phone),
       Appointment.type.in_(service_names)
   ).all()
   
   if not appointments:
       raise HTTPException(
           status_code=404,
           detail="No appointments found for this phone number with your business"
       )
   
   return appointments

@router.get("/appointments/stats/{date}")
def get_appointments_stats(
    date: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(business_owner_required)
):
    try:
        # Get the business owner
        user = db.query(User).filter(User.username == current_user["sub"]).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        # Convert date string to date object
        target_date = datetime.strptime(date, "%m-%d-%Y").date()
        
        # Get owner's services
        services = db.query(Service).filter(Service.owner_id == user.id).all()
        service_names = [service.name for service in services]

        # Get appointments for the date that match owner's services
        appointments = db.query(Appointment).filter(
            text("DATE(date) = DATE(:target_date)"),
            Appointment.type.in_(service_names)
        ).params(target_date=target_date).all()

        if not appointments:
            raise HTTPException(
                status_code=404,
                detail=f"No appointments found for {target_date.strftime('%m/%d/%Y')}"
            )

        # Calculate statistics by service type
        service_stats = {}
        total_revenue = 0

        for appointment in appointments:
            if appointment.type not in service_stats:
                service_stats[appointment.type] = {
                    'count': 0,
                    'revenue': 0
                }
            
            service_stats[appointment.type]['count'] += 1
            service_stats[appointment.type]['revenue'] += appointment.cost
            total_revenue += appointment.cost

        return {
            "date": target_date.strftime("%m/%d/%Y"),
            "total_revenue": total_revenue,
            "services": service_stats
        }

    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Invalid date format. Please use MM-DD-YYYY format.")


@router.get("/my-appointments", response_model=List[AppointmentResponse])
async def get_business_appointments(
    db: Session = Depends(get_db),
    current_user: dict = Depends(business_owner_required)
):
    # Get the business owner's info
    user = db.query(User).filter(User.username == current_user["sub"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Get all appointments for this business owner's services
    appointments = db.query(Appointment).join(
        Service, 
        and_(
            Service.name == Appointment.type,
            Service.owner_id == user.id
        )
    ).order_by(Appointment.date.desc()).all()

    return appointments

def get_month_date_range():
    today = datetime.now()
    start_date = today.replace(day=1)
    if today.month == 12:
        end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    return start_date, end_date

@router.get("/business/appointments/export")
async def export_appointments_to_excel(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "business_owner":
        raise HTTPException(status_code=403, detail="Only business owners can export appointments")

    # Get appointments for the current month
    start_date, end_date = get_month_date_range()
    
    appointments = (
        db.query(
            Appointment,
            Service.name.label("service_name"),
            Service.price.label("service_price"),
            User.first_name.label("customer_first_name"),
            User.last_name.label("customer_last_name"),
            User.phone.label("customer_phone")
        )
        .join(Service, Appointment.title == Service.name)
        .join(User, Appointment.customer_id == User.id)
        .filter(Appointment.business_id == current_user.id)
        .filter(Appointment.date >= start_date)
        .filter(Appointment.date <= end_date)
        .all()
    )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments Report"

    # Styles
    header_style = NamedStyle(name="header_style")
    header_style.fill = PatternFill("solid", fgColor="1F4E78")
    header_style.font = Font(color="FFFFFF", bold=True, size=12)
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    data_style = NamedStyle(name="data_style")
    data_style.font = Font(size=11)
    data_style.alignment = Alignment(horizontal="left", vertical="center")
    data_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{current_user.business_name} - Monthly Appointments Report"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Add date range
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Report Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Date",
        "Time",
        "Service",
        "Customer Name",
        "Phone",
        "Duration (min)",
        "Revenue",
        "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.style = header_style

    # Add data
    total_revenue = 0
    services_count = {}

    for row_idx, (appt, service_name, service_price, first_name, last_name, phone) in enumerate(appointments, start=5):
        # Track statistics
        total_revenue += service_price
        services_count[service_name] = services_count.get(service_name, 0) + 1

        # Format data
        date = appt.date.strftime("%Y-%m-%d")
        time = appt.start_time.strftime("%H:%M")
        customer_name = f"{first_name} {last_name}"
        
        # Add row data
        row_data = [
            date,
            time,
            service_name,
            customer_name,
            phone,
            appt.duration,
            f"${service_price:,.2f}",
            "Completed" if appt.date < datetime.now().date() else "Scheduled"
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.style = data_style

    # Add summary section
    summary_row = len(appointments) + 6
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    ws[f'A{summary_row}'] = "Summary"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    ws[f'A{summary_row}'].fill = PatternFill("solid", fgColor="E7EDF5")

    # Summary data
    average_revenue = total_revenue / len(appointments) if appointments else 0
    most_popular_service = max(services_count.items(), key=lambda x: x[1])[0] if services_count else "N/A"

    summary_data = [
        ("Total Appointments", len(appointments)),
        ("Total Revenue", f"${total_revenue:,.2f}"),
        ("Average Revenue", f"${average_revenue:,.2f}"),
        ("Most Popular Service", most_popular_service)
    ]

    for i, (label, value) in enumerate(summary_data):
        row = summary_row + i + 1
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)

    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the Excel file
    headers = {
        'Content-Disposition': 'attachment; filename=appointments_report.xlsx'
    }
    
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )